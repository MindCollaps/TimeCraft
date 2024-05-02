from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
import datetime
import requests
import argparse
import glob
import json
import re


class excel_parser():
    def __init__(self, file_path):
        self.TABLE_HEIGHT = 9
        self.TABLE_WIDTH = 8
        self.CELLS_BETWEEN_TABLES = 1
        self.TABLE_START_OFFSET = 4

        self.weekdays = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"]

        self.holidays = ["Tag der Deutschen Einheit", "Reformationstag", "Allerheiligen", "Buß- und Bettag",
                         "Weihnachten", "Neujahr", "Heilige Drei Könige", "Karfreitag", "Ostermontag", "Tag der Arbeit",
                         "Christi Himmelfahrt", "Pfingstmontag", "Fronleichnam"]

        self.times = ["8:00-9:00", "9:00-12:15", "12:15-13:00", "13:00-14:30", "14:45-16:15", "16:30-18:00",
                      "18:15-19:45"]

        self.SemesterGroups = ["S1", "S2", "A1", "A2"]
        self.SpecialGroups = ["S1-1", "S1-2", "S2-1", "S2-2", "A1-1", "A1-2", "A2-1", "A2-2"]

        self.KnownParallelLessons = ["EAC", "OOP", "Gruppensprecher-Runde", "Mathetutorium", "ITIL", "Krypto", "ITS-M",]
        self.KnownEvents = ["Science Winter", "Science Summer", "Science Spring", "Semesterinfo", "Ersti-Begrüßung",
                            "Gruppensprecher-Runde", "Präsentation PR1 Zweitversuch"]
        
        self.KnownLecturer = {
            "4. Semester": {
                "S1": {
                    "Stat II": "Neumann-Brosig",
                    "AZ2": "Neumann-Brosig",
                    "PF": "Ahlers",
                    "GCC": "Werner",
                    "Projekt": "Schmidt",
                    "ACC": "Birzer",
                    "Krypto": "Neubauer",
                    "ITS-M": "Peine-Paulsen",
                    "ITIL": "Schaper",
                    "ITAA": "Ibrahim",
                    "SiS": "Stephanus",
                    "ITG/C": "Riebandt",
                },
                "S2": {
                    "Stat II": "Neumann-Brosig",
                    "AZ2": "Neumann-Brosig",
                    "PF": "Ahlers",
                    "GCC": "Werner",
                    "Projekt": "Lobachev",
                    "ACC": "Birzer",
                    "Krypto": "Neubauer",
                    "ITS-M": "Peine-Paulsen",
                    "ITIL": "Schaper",
                    "ITAA": "Ibrahim",
                    "SiS": "Stephanus",
                    "ITG/C": "Riebandt",
                }
            }
        }

        self.file_path = file_path
        self.matrix = None

        self.current_day_counter = 0
        self.current_time_counter = None

        self.__load__(self.file_path)
        self.convert_to_matrix(self.wb)

    def __load__(self, file_path):
        self.wb = load_workbook(filename=file_path)
        self.ws = self.wb.active

        # unmerge all cells and fill the empty cells with the top value
        merged_cell_values = self.create_merged_cell_merged_cell_values(self.ws)
        cell_group_list = merged_cell_values.keys()
        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            self.ws.unmerge_cells(str(cell_group))
            for row in self.ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = merged_cell_values[cell_group]

        self.rows = list(self.ws.rows)

    def create_merged_cell_merged_cell_values(self, sheet) -> dict:
        """
        :param sheet:
        :return: the key-value pairs (dict) of merged cell and top value

        source: https://dasiyql.medium.com/unmerge-excel-cells-in-python-and-populate-the-top-cell-value-c567c8490a11
        """
        merged_merged_cell_values = {}
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            if min_col == max_col:
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                merged_merged_cell_values[str(cell_group)] = top_left_cell_value
        return merged_merged_cell_values

    def convert_to_matrix(self, wb):
        # get the active worksheet
        ws = wb.active

        # get the max row count
        max_row = ws.max_row

        # get the max column count
        max_column = ws.max_column

        # iterate over all cells
        # iterate over all rows
        # save all data
        data = {}
        for i in range(1, max_row + 1):
            for j in range(1, max_column + 1):
                # get particular cell value
                cell_obj = ws.cell(row=i, column=j)

                letter = self.getCoordinateLetter(cell_obj)

                # save the letter
                if letter not in data:
                    data[letter] = {}

                # save the value
                data[letter][cell_obj.row] = cell_obj.value

        self.matrix = data

    def extract_tables(self, rows):
        tables = []

        # header values
        study_subject = None            # dIT2022
        self.semester_group = None           # S1, S2, A1, A2
        semester = None                 # WS2023/2024
        self.semester_year = None            # 3. Semester
        last_changed = None             # Stand: 06.11.2023

        # start values
        row_counter = 1                 # TABLE_HEIGHT --> from 1 to 9
        cell_counter = 1                # TABLE_WIDTH --> from 1 to 8
        header_parsed = False

        table = {
            "days": {},                 # contains the days of the week
            "calendarweek": None,       # contains the calendar week
            "start_row": None,          # the start row of the table, like 5
            "end_row": None,            # the end row of the table, like 13 --> table height is 8
        }

        skip_next_row = False           # skip the next row, because it is the space between two tables
        skipped_rows = 0                # count the skipped rows

        for row in rows:
            cell_counter = 1

            if not skip_next_row:
                for cell in row:
                    # skip all empty cells after the table
                    if cell_counter > self.TABLE_WIDTH:
                        continue

                    if not header_parsed:
                        if row_counter == self.TABLE_START_OFFSET + 1:
                            header_parsed = True

                            # save the header at the first position
                            tables.append({
                                "study_subject": study_subject,
                                "semester_group": self.semester_group,
                                "semester": semester,
                                "semester_year": self.semester_year,
                                "last_changed": last_changed,
                            })

                    if not header_parsed:
                        # empty cells in the header are not important
                        if cell.value != None:

                            if self.containsYear(cell.value) and not ("WS" in cell.value or "SS" in cell.value or "SoSe" in cell.value or "WiSe" in cell.value) and not self.containsLastChanged(cell.value):
                                study_subject, self.semester_group = self.extractSemesterGroup(cell.value)

                            elif ("WS" in cell.value or "SS" in cell.value or "SoSe" in cell.value or "WiSe" in cell.value):
                                semester = cell.value

                            elif self.containsSemester(cell.value):
                                self.semester_year = cell.value

                            elif self.containsLastChanged(cell.value):
                                last_changed = cell.value

                    else:
                        # finally parse the time table

                        # to parse the empty cells later
                        # because they could be free lessons
                        if cell.value == None:
                            cell_content = ""
                        else:
                            cell_content = cell.value

                        # get the real value
                        # this has to happen before the other checks
                        if not self.isNumber(cell_content) and not self.isDatetime(cell_content) and "=" in cell_content:
                            letter, index, increaseBy = self.extractCoordinates(cell_content)
                            actual_value = self.matrix[letter][index]

                            if self.isDatetime(actual_value):
                                actual_value = actual_value + datetime.timedelta(days=int(increaseBy))
                            elif self.isNumber(actual_value):
                                actual_value = actual_value + int(increaseBy)
                            else:
                                print("Error: Unknown type")

                            # save the new value
                            cell_content = actual_value
                            self.matrix[self.getCoordinateLetter(cell)][cell.row] = actual_value

                            # do not continue, because we want to check the new value
                            # continue

                        # fix text
                        cell_content = self.fixText(cell_content)
                        cell.value = cell_content

                        if self.isWeekday(cell_content):
                            if table["days"] == {}:
                                table["days"] = []
                            table["days"].append({})
                            table["days"][self.current_day_counter] = {}

                            self.updateDayCounter()

                        elif self.isNumber(cell_content):
                            # check if we are in the second row of the table
                            if (table["start_row"] + 1) == row_counter:
                                table["calendarweek"] = cell_content

                        elif self.isDatetime(cell_content):

                            table["days"][self.current_day_counter]["date"] = cell_content
                            self.updateDayCounter()

                        elif self.containsTime(cell_content):
                            # save the time for each day
                            for day in range (0, 6):
                                if "lessons" not in table["days"][day]:
                                    table["days"][day]["lessons"] = []

                            if self.current_time_counter == None:
                                self.current_time_counter = 0
                            else:
                                self.current_time_counter += 1

                        # there is no relevant content before cell_counter 2
                        elif cell_counter > 2 and not "KW" in cell_content and table["start_row"] != None and table["end_row"] != None:
                            # check if the cell is a parallel lesson
                            if self.isParallelLesson(cell_content):
                                lessons = self.getParallelLessons(cell_content)
                            else:
                                lessons = [cell_content]

                            # iterate through all lessons
                            # lesson contains the modified cellcontent
                            for lesson in lessons:
                                current_time = self.getTimesByIndex(self.current_time_counter)
                                current_day = self.current_day_counter
                                index_counter = len(table["days"][current_day]["lessons"])

                                table["days"][current_day]["lessons"].append({})

                                lesson = lesson.replace("*", "").replace("\n", " ").replace("online ", "").strip()
                                if self.isCustomTime(lesson):
                                    time = self.getCustomTime(lesson)
                                    table["days"][current_day]["lessons"][index_counter]["time"] = self.parseTime(time, current_time)

                                    # delete the custom time from the cell content
                                    replace_strings = self.getCustomTime(lesson, True)
                                    for string in replace_strings:
                                        lesson = lesson.replace(string, "")

                                    # remove \n and spaces
                                    lesson = lesson.strip()
                                else:
                                    table["days"][current_day]["lessons"][index_counter]["time"] = current_time

                                if self.isHoliday(lesson):
                                    text = f"Feiertag: {lesson}"
                                    table["days"][current_day]["lessons"][index_counter]["isHoliday"] = True
                                elif lesson == "":
                                    text = "no lesson --> freetime"
                                    table["days"][current_day]["lessons"][index_counter]["isHoliday"] = False
                                else:
                                    table["days"][current_day]["lessons"][index_counter]["isHoliday"] = False
                                    text = lesson

                                skip = False

                                if index_counter != 0:
                                    # remove the lecturer name from the lesson name
                                    lecturer = self.getLecturer(text)
                                    cleaned_text = text.replace(f" ({lecturer})", "").strip()

                                    if self.isExtendedLesson(cleaned_text, table["days"][current_day]["lessons"][index_counter - 1]):
                                        current_lession = table["days"][current_day]["lessons"][index_counter]
                                        old_lession = table["days"][current_day]["lessons"][index_counter - 1]

                                        table["days"][current_day]["lessons"][index_counter - 1][
                                            "time"] = self.getExtendedLessonTime(current_lession, old_lession)

                                        # delete the lesson from the array
                                        table["days"][current_day]["lessons"].pop(index_counter)

                                        # don't add a new lesson
                                        skip = True

                                if not skip:
                                    table["days"][current_day]["lessons"][index_counter]["name"] = text

                                    if self.isOnline(cell):
                                        table["days"][current_day]["lessons"][index_counter]["isOnline"] = True
                                    else:
                                        table["days"][current_day]["lessons"][index_counter]["isOnline"] = False

                                    if lesson != "" and self.isReExamination(cell):
                                        table["days"][current_day]["lessons"][index_counter]["isReExamination"] = True
                                    else:
                                        table["days"][current_day]["lessons"][index_counter]["isReExamination"] = False

                                    if lesson != "" and self.isExam(cell):
                                        table["days"][current_day]["lessons"][index_counter]["isExam"] = True
                                    else:
                                        table["days"][current_day]["lessons"][index_counter]["isExam"] = False

                                    if lesson != "" and self.wasCanceled(cell):
                                        table["days"][current_day]["lessons"][index_counter]["wasCanceled"] = True
                                    else:
                                        table["days"][current_day]["lessons"][index_counter]["wasCanceled"] = False

                                    if lesson != "" and self.wasMoved(cell):
                                        table["days"][current_day]["lessons"][index_counter]["wasMoved"] = True
                                    else:
                                        table["days"][current_day]["lessons"][index_counter]["wasMoved"] = False

                                    lecturer = self.getLecturer(lesson)
                                    table["days"][current_day]["lessons"][index_counter]["lecturer"] = lecturer

                                    # remove the lecturer name from the lesson name
                                    if lecturer != None:
                                        lesson_name = lesson.replace(f" ({lecturer})", "").strip()
                                        table["days"][current_day]["lessons"][index_counter]["name"] = lesson_name

                                    if lesson != "" and self.isEvent(lesson, cell):
                                        table["days"][current_day]["lessons"][index_counter]["isEvent"] = True
                                    else:
                                        table["days"][current_day]["lessons"][index_counter]["isEvent"] = False

                                    # table["days"][current_day]["lessons"][index_counter]["room"] = self.getRoom(
                                    #     table["days"][current_day]["lessons"][index_counter]["name"],
                                    #     cell,
                                    #     table["days"][current_day]["date"],
                                    #     study_subject,
                                    #     self.semester_group,
                                    #     lecturer
                                    # )

                            self.updateDayCounter()

                        elif "KW" in cell_content:
                            # calculate the start and end row of the table
                            table["start_row"] = row_counter
                            table["end_row"] = row_counter + self.TABLE_HEIGHT - 1

                        # check if we need a new table for a new week
                        # after the last row / cell was parsed
                        # row_counter > self.TABLE_START_OFFSET --> skip the first table
                        if ((row_counter - skipped_rows) - self.TABLE_START_OFFSET) % self.TABLE_HEIGHT == 0 and cell_counter == 8 and row_counter > self.TABLE_START_OFFSET:
                            # new table
                            if table["days"] != {}:
                                tables.append(table)

                            table = {
                                "days": {},             # contains contains the days of the week
                                "calendarweek": None,   # contains the calendar week
                                "start_row": None,      # the start row of the table, like 5
                                "end_row": None,        # the end row of the table, like 13 --> table height is 8
                            }

                            self.current_day_counter = 0
                            self.current_time_counter = None

                            # skip the next row, because it is the space between two tables
                            skip_next_row = True

                    cell_counter += 1

            else:
                # check if the next row is actually the start of a new table
                # if not, skip the row
                next_start_row = row_counter + 1
                row_value = self.rows[next_start_row - 1][1]
                if row_value.value == "KW":
                    # row skipped successfully
                    skip_next_row = False
                else:
                    print("Detected invalid space between two tables. Skipping this row ...")

                skipped_rows += 1

            row_counter += 1

        return tables

    # "contains" methods
    def containsYear(self, value):
        match = re.search(r'\d{2,4}', value)
        return match != None

    def containsSemester(self, value):
        # https://regex101.com/r/6aET8G/1
        match = re.search(r'\d{1,2}.?\sSemester', value)
        return match != None

    def containsLastChanged(self, value):
        # https://regex101.com/r/WJNP5m/1
        match = re.search(r'Stand:\s\d{2}.\d{2}.\d{4}', value)
        return match != None

    def containsTime(self, value):
        # https://regex101.com/r/xSKgS3/1
        match = re.search(r'^\d{1,2}:\d{1,2}-\d{1,2}:\d{1,2}$', value)
        return match != None

    # type check methods
    def isNumber(self, value):
        return type(value) == int or type(value) == float

    def isDatetime(self, value):
        return type(value) == datetime.datetime

    def isWeekday(self, value):
        return value in self.weekdays

    def isHoliday(self, value):
        return value in self.holidays

    def isOnline(self, cell):
        hexcolor = self.getFillColor(cell)
        return hexcolor == "FFFFFF00" or "online" in cell.value

    def isBold(self, cell):
        return cell.font.b

    def isReExamination(self, cell):
        return "NKL" in cell.value #and self.isBold(cell)

    def isExam(self, cell):
        return "Klausur" in cell.value and self.isBold(cell) and "NKL" not in cell.value

    def isStrikeThrough(self, cell):
        return cell.font.strike

    def isCustomTime(self, value):
        # https://regex101.com/r/DzObO7/1
        match = re.search(
            r'[0-9]{1,2}:[0-9]{2}-[0-9]{1,2}:[0-9]{2}|Beginn [0-9]{1,2}:[0-9]{2}|ab [0-9]{1,2}:[0-9]{2}|bis [0-9]{1,2}:[0-9]{2}|[0-9]{1,2}:[0-9]{2}',
            value)
        return match != None

    def isExtendedLesson(self, value, last_lesson):
        # Sprachenzentrum needs to be seperate --> different groups
        return last_lesson["name"] == value and last_lesson["name"] not in ["Sprachenzentrum"]

    def isParallelLesson(self, value):
        # check if there are two teachers / two lesson names / semester group names like S1, S2, A1, A2, ....
        # Example: "EAC S2-1 (Arnold)\nEAC S2-2 (Schmidt)"
        # Example: "EAC S2-1 (Arnold) / EAC S2-2 (Schmidt)"

        SemesterGroupMatched = any(item in value for item in self.SemesterGroups)
        KnownParallelLessonsMatched = any(item in value for item in self.KnownParallelLessons)
        # TODO: check if there are two teachers

        return SemesterGroupMatched or KnownParallelLessonsMatched

    def isEvent(self, value, cell):
        match = any(item in value for item in self.KnownEvents)

        if match or (self.getLecturer(value) == None and not self.isReExamination(cell) and not self.isExam(cell)
                     and not self.isHoliday(value) and not self.isSpeechLesson(value)
                     and not value in self.KnownParallelLessons
                     and not self.isReExamination(cell) and not self.isExam(cell)
                     ):
            return True
        else:
            return False

    def isSpeechLesson(self, value):
        return "Sprachenzentrum" in value or "Sprachenzentrum" == value

    # was methods
    def wasCanceled(self, cell):
        return self.isStrikeThrough(cell) # and self.getFontColor(cell) == "FFFF0000"  # red

    def wasMoved(self, cell):
        return (not self.isStrikeThrough(cell) and self.getFontColor(cell) == "FFFF0000")  # red

    # getter methods
    def getWeekdayByIndex(self, index):
        return self.weekdays[index]

    def getExtendedLessonTime(self, lesson, last_lesson):
        return last_lesson["time"].split("-")[0] + "-" + lesson["time"].split("-")[1]

    def getCustomTime(self, value, returnAll=False):
        custom_start_end = re.search(r'[0-9]{1,2}:[0-9]{2}-[0-9]{1,2}:[0-9]{2}', value)  # 8:00-9:00
        custom_start = re.findall(r'[0-9]{1,2}:[0-9]{2}', value)  # Semesterinfo 8:00
        begin = re.findall(r'Beginn [0-9]{1,2}:[0-9]{2}|ab [0-9]{1,2}:[0-9]{2}', value)  # Beginn 8:00
        til = re.findall(r'bis [0-9]{1,2}:[0-9]{2}', value)  # bis 9:00

        if custom_start_end != None:
            if returnAll:
                return [custom_start_end.group(0)]
            else:
                return custom_start_end.group(0)
        elif begin != []:
            if returnAll:
                return begin
            elif len(begin) > 1:
                return begin[1]
            else:
                return begin[0]
        elif til != []:
            if returnAll:
                return til
            elif len(til) > 1:
                return til[1]
            else:
                return til[0]
        elif custom_start != []:
            if returnAll:
                return custom_start
            elif len(custom_start) < 1:
                return custom_start[1]
            else:
                return custom_start[0]
        else:
            if returnAll:
                return []
            else:
                return None

    def getTimesByIndex(self, index):
        return self.times[index]

    def getCoordinateLetter(self, cell_obj):
        # check if the cell object has a letter
        # if not, get it from the coordinate

        if hasattr(cell_obj, "column_letter"):
            letter = cell_obj.column_letter
        else:
            letter = cell_obj.coordinate[0]

        return letter

    def getFillColor(self, cell):
        # hexadecimal value of the color
        return cell.fill.start_color.index

    def getFontColor(self, cell):
        if cell.font.color != None:
            return cell.font.color.index
        else:
            return None

    def getLecturer(self, value):
        # https://regex101.com/r/mUYVjM/1
        match = re.search(r'\(([A-Za-z-.]*)\)', value)
        if match:
            lecturer = match.group(1)
        else:
            lecturer = self.KnownLecturer.get(self.semester_year, {}).get(self.semester_group, {}).get(value, None)

        return lecturer

    def getParallelLessons(self, value):
        if "/" in value:
            return value.split("/")
        else:
            return value.rsplit("\n")

    def fixText(self, text):
        # fix common typing mistakes in the timetable
        map = {
            "LingAlg": "LinAlg",
            "Zeitversuch": "Zweitversuch",
            "Satistik": "Statistik",
            "IST-M": "ITS-M",
            "Stat2": "Stat II",
        }

        # only fix strings and skip integers
        if isinstance(text, str):
            for key, value in map.items():
                text = text.replace(key, value)

        return text

    def getRoom(self, value, cell, date, study_subject, semester_group, lecturer):
        date = date.strftime("%Y-%m-%d")
        r = requests.get(f"https://digitalroomplan-api.azurewebsites.net/plans/{date}")

        if value == "Sprachenzentrum":
            # not enough infos to differentiate the different Sprachzentren
            return "Not possible as for now"

        if value in self.KnownEvents or "Feiertag" in value or value == "no lesson --> freetime":
            return None

        if value == "":
            return None

        # look for special groups
        if any(item in value for item in self.SpecialGroups):
            # https://regex101.com/r/eFCdkm/1
            special_group = re.search(r'[AS]{1,2}[0-9]{1,2}-[0-9]{1,2}', value).group(0)
            value = value.replace(f" {special_group}", "").strip()
        else:
            special_group = ""

        # strip numbers and roman numbers
        # because 1 and I do not match
        value = re.sub(r'\s?[0-9]+', '', value)
        value = re.sub(r'\s?[IVX]+', '', value)

        if r.status_code == 200:
            data = r.json()

            found = False
            rooms = []
            for item in data["rooms"]:
                if item["occupancies"] != []:
                    for occupancy in item["occupancies"]:
                        tags = occupancy["tags"]

                        # empty text will match also
                        if special_group != '' and special_group not in tags:
                            continue

                        if value not in occupancy["text"]:
                            continue

                        if study_subject not in occupancy["text"]:
                            continue

                        if semester_group not in occupancy["text"]:
                            continue

                        if lecturer not in occupancy["text"]:
                            continue

                        found = True
                        rooms.append(item["identifier"])

            if found:
                return rooms
            else:
                return None
        else:
            return None

    # update methods
    def updateDayCounter(self):
        if self.current_day_counter != 0 and self.current_day_counter % 5 == 0:
            self.current_day_counter = 0
        else:
            self.current_day_counter += 1

    # extract methods
    def extractSemesterGroup(self, value):
        # https://regex101.com/r/nLR5HB/1
        match = re.search(r'([a-zA-z]{3,5}\d{2,4})\s?([a-zA-Z]\d)?', value)
        return match.group(1), match.group(2)

    def extractCoordinates(self, text):
        """ extract the letter, the index and the content from a string
            like =C6+1
        """
        text = text.replace("=", "")

        letter = text[0]
        index = text[1:]

        # some cells have a + in the index
        # like =C6+1
        content = None

        if "+" in index:
            temp = index.split("+")
            index = temp[0]
            content = temp[1]

        return letter, int(index), content

    def parseTime(self, value, old_time):
        if value.startswith("Beginn") or value.startswith("ab"):
            # get the time (like 8:00)
            time = value.split(" ")[1]

            # get end of old time (like 9:00)
            end = old_time.split("-")[1]

            # return the new time (like 8:00-9:00)
            return f"{time}-{end}"
        elif value.startswith("bis"):
            # get the time (like 9:00)
            time = value.split(" ")[1]

            # get start of old time (like 8:00)
            start = old_time.split("-")[0]

            # return the new time (like 8:00-9:00)
            return f"{start}-{time}"
        elif not "-" in value and not "Beginn" in value and not "ab" in value and not "bis" in value:
            return f"{value}-{old_time.split("-")[1]}"  # start time is new
        else:
            return value


if __name__ == "__main__":
    argparser = argparse.ArgumentParser(description="Extract the timetable from the excel file")
    argparser.add_argument("-f", "--file", help="The excel file to parse")

    args = argparser.parse_args()
    files = []
    
    if args.file:
        files.append(args.file)
    else:
        # get all *.xlsx files from the current directory
        files = [file for file in glob.glob("*.xlsx")]
        
    for file in files:
        print(f"Parsing {file}")
        parser = excel_parser(file)
        tables = parser.extract_tables(parser.rows)

        with open(file.replace(".xlsx", ".json"), 'w', encoding='utf-8') as f:
            json.dump(tables, f, ensure_ascii=False, indent=4, default=str)

    print("Done")

# TODO: support for 
# EAC S2-1 (Arnold) / (durchgestrichen)
# EAC S2-2 (Schmidt)

# TODO: correct freetime
#        --> 8:00-9:00 free and then a lesson from 10.45-12.15
# TODO: parse Science Winter outside of a table
# TODO: parse other timetables 
#       --> S1, A1 etc
